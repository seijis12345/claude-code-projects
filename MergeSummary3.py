# =============================================================================
# MergeSummary3.py
# 3つのSummaryExcelファイルのグラフ系列をマージするプログラム
# MergeSummary.py を3ファイル対応に拡張したバージョン
# =============================================================================

import copy
import shutil
import os
import re
import openpyxl

# =============================================================================
# ★★★ 設定エリア ★★★
# =============================================================================

BASE_DIR = r"C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\40365"

# ファイル1（Heatup）
FILE1_PATH = os.path.join(BASE_DIR, r"Cold_Hot_Eval\40365 Summary_heatup_XVDAC.xlsx")
FILE1_SUFFIX = "Heatup"

# ファイル2（1st measurement）
FILE2_PATH = os.path.join(BASE_DIR, r"Repeat_40365\40365 Summary_1st measurement.xlsx")
FILE2_SUFFIX = "1st"

# ファイル3（2nd measurement）
FILE3_PATH = os.path.join(BASE_DIR, r"Repeat_40365\40365 Summary_2nd measurement.xlsx")
FILE3_SUFFIX = "2nd"

# 出力ファイルの保存先フォルダ（空欄の場合はFILE1_PATHと同じフォルダ）
OUTPUT_DIR = ""

# =============================================================================
# ★★★ 設定エリア ここまで ★★★
# =============================================================================

FILE2_SHEET_PREFIX = "_B_"
FILE3_SHEET_PREFIX = "_C_"


def get_series_title(ser):
    try:
        if ser.title is None:
            return None
        if hasattr(ser.title, 'v') and ser.title.v is not None:
            return ser.title.v
        if hasattr(ser.title, 'strRef') and ser.title.strRef:
            return ser.title.strRef.f
        return str(ser.title)
    except Exception:
        return None


def set_series_title(ser, new_title):
    if ser.title is not None and hasattr(ser.title, 'v') and ser.title.v is not None:
        ser.title.v = new_title
    elif ser.title is not None and hasattr(ser.title, 'strRef') and ser.title.strRef:
        ser.title.strRef.f = None
        ser.title.v = new_title
    else:
        from openpyxl.chart.series import SeriesLabel
        ser.title = SeriesLabel(v=new_title)


def copy_sheet_values(ws_src, ws_dst):
    """data_only=Trueで読んだシートの計算値をコピーする"""
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value is not None:
                ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)


def update_series_sheet_refs(ser, sheet_name_map):
    """系列内のシート参照をsheet_name_mapに従って書き換え、キャッシュをクリアする"""
    def remap(formula):
        if not formula:
            return formula
        for old_name, new_name in sheet_name_map.items():
            formula = formula.replace(f"'{old_name}'!", f"'{new_name}'!")
            formula = re.sub(
                rf"(?<!['\w])({re.escape(old_name)})!",
                f"'{new_name}'!",
                formula
            )
        return formula

    for attr in ['xVal', 'yVal', 'val']:
        ds = getattr(ser, attr, None)
        if ds and hasattr(ds, 'numRef') and ds.numRef and ds.numRef.f:
            ds.numRef.f = remap(ds.numRef.f)
            ds.numRef.numCache = None


def collect_ref_sheets(wb_charts, chart_sheets):
    """チャートシートから参照データシート名を収集する"""
    ref_sheets = set()
    for sname in chart_sheets:
        ws = wb_charts[sname]
        for chart in ws._charts:
            for ser in chart.series:
                for attr in ['xVal', 'yVal', 'val']:
                    ds = getattr(ser, attr, None)
                    if ds and hasattr(ds, 'numRef') and ds.numRef and ds.numRef.f:
                        raw = ds.numRef.f.split('!')[0].strip("'")
                        ref_sheets.add(raw)
    return ref_sheets


def copy_data_sheets(wb_data, wb_merged, ref_sheets, prefix, label):
    """データシートをマージファイルにコピーし、シート名マップを返す"""
    sheet_name_map = {}
    for sname in sorted(ref_sheets):
        if sname not in wb_data.sheetnames:
            print(f"  警告: '{sname}' が{label}に見つかりません。スキップします。")
            continue
        new_name = f"{prefix}{sname}"[:31]
        # 同名シートが既に存在する場合はスキップ
        if new_name in wb_merged.sheetnames:
            sheet_name_map[sname] = new_name
            continue
        sheet_name_map[sname] = new_name
        ws_src = wb_data[sname]
        ws_dst = wb_merged.create_sheet(title=new_name)
        copy_sheet_values(ws_src, ws_dst)
        count = sum(1 for row in ws_dst.iter_rows() for c in row if c.value is not None)
        print(f"  データシートコピー: '{sname}' → '{new_name}' ({count} セル) [{label}]")
    return sheet_name_map


def merge_summary3(file1_path, file1_suffix,
                   file2_path, file2_suffix,
                   file3_path, file3_suffix,
                   output_dir=""):
    """3つのExcelファイルのグラフ系列をマージする"""

    base_dir = output_dir if output_dir else os.path.dirname(file1_path)
    f1_name = os.path.splitext(os.path.basename(file1_path))[0]
    output_filename = f"Compare_{f1_name}.xlsx"
    output_path = os.path.join(base_dir, output_filename)

    print(f"マージ開始...")
    print(f"  File1 ({file1_suffix}): {os.path.basename(file1_path)}")
    print(f"  File2 ({file2_suffix}): {os.path.basename(file2_path)}")
    print(f"  File3 ({file3_suffix}): {os.path.basename(file3_path)}")
    print(f"  出力: {output_filename}")
    print()

    # File1をベースとしてコピー
    shutil.copy2(file1_path, output_path)
    wb_merged = openpyxl.load_workbook(output_path)

    # File2: チャート参照用 / データ取得用
    wb2_charts = openpyxl.load_workbook(file2_path)
    wb2_data   = openpyxl.load_workbook(file2_path, data_only=True)

    # File3: チャート参照用 / データ取得用
    wb3_charts = openpyxl.load_workbook(file3_path)
    wb3_data   = openpyxl.load_workbook(file3_path, data_only=True)

    # グラフのある共通シートを特定
    chart_sheets_2 = [s for s in wb_merged.sheetnames
                      if s in wb2_charts.sheetnames and wb_merged[s]._charts]
    chart_sheets_3 = [s for s in wb_merged.sheetnames
                      if s in wb3_charts.sheetnames and wb_merged[s]._charts]

    # --- Step 1: File2のデータシートをコピー ---
    file2_ref_sheets = collect_ref_sheets(wb2_charts, chart_sheets_2)
    sheet_name_map2 = copy_data_sheets(wb2_data, wb_merged, file2_ref_sheets, FILE2_SHEET_PREFIX, "File2")

    # --- Step 2: File3のデータシートをコピー ---
    file3_ref_sheets = collect_ref_sheets(wb3_charts, chart_sheets_3)
    sheet_name_map3 = copy_data_sheets(wb3_data, wb_merged, file3_ref_sheets, FILE3_SHEET_PREFIX, "File3")

    print()

    # --- Step 3: チャート系列のマージ ---
    total_charts = 0
    total_series_renamed = 0
    total_series_added_2 = 0
    total_series_added_3 = 0

    # File2とFile3の両方でチャートのある全シートを対象にする
    all_chart_sheets = sorted(set(chart_sheets_2) | set(chart_sheets_3))

    for sheet_name in all_chart_sheets:
        ws_merged = wb_merged[sheet_name]
        print(f"  シート '{sheet_name}': {len(ws_merged._charts)} チャート処理中...")

        for chart_idx, chart_merged in enumerate(ws_merged._charts):
            total_charts += 1

            # File1系列名に suffix1 を付加
            for ser in chart_merged.series:
                original_title = get_series_title(ser)
                if original_title is not None:
                    set_series_title(ser, f"{original_title} {file1_suffix}")
                    total_series_renamed += 1

            # File2系列を追加
            if sheet_name in chart_sheets_2:
                ws2 = wb2_charts[sheet_name]
                if chart_idx < len(ws2._charts):
                    for ser2 in ws2._charts[chart_idx].series:
                        ser_copy = copy.deepcopy(ser2)
                        update_series_sheet_refs(ser_copy, sheet_name_map2)
                        orig = get_series_title(ser_copy)
                        if orig is not None:
                            set_series_title(ser_copy, f"{orig} {file2_suffix}")
                        chart_merged.series.append(ser_copy)
                        total_series_added_2 += 1

            # File3系列を追加
            if sheet_name in chart_sheets_3:
                ws3 = wb3_charts[sheet_name]
                if chart_idx < len(ws3._charts):
                    for ser3 in ws3._charts[chart_idx].series:
                        ser_copy = copy.deepcopy(ser3)
                        update_series_sheet_refs(ser_copy, sheet_name_map3)
                        orig = get_series_title(ser_copy)
                        if orig is not None:
                            set_series_title(ser_copy, f"{orig} {file3_suffix}")
                        chart_merged.series.append(ser_copy)
                        total_series_added_3 += 1

    wb_merged.save(output_path)
    wb2_charts.close()
    wb2_data.close()
    wb3_charts.close()
    wb3_data.close()

    print()
    print(f"マージ完了!")
    print(f"  処理チャート数        : {total_charts}")
    print(f"  リネーム系列数        : {total_series_renamed} ({file1_suffix}付加)")
    print(f"  追加系列数 (File2)    : {total_series_added_2} ({file2_suffix}付加)")
    print(f"  追加系列数 (File3)    : {total_series_added_3} ({file3_suffix}付加)")
    print(f"  コピーシート数        : {len(sheet_name_map2) + len(sheet_name_map3)}")
    print(f"  保存先                : {output_path}")

    return output_path


if __name__ == "__main__":
    merge_summary3(
        file1_path=FILE1_PATH,
        file1_suffix=FILE1_SUFFIX,
        file2_path=FILE2_PATH,
        file2_suffix=FILE2_SUFFIX,
        file3_path=FILE3_PATH,
        file3_suffix=FILE3_SUFFIX,
        output_dir=OUTPUT_DIR,
    )
