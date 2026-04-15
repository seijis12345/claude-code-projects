# =============================================================================
# MergeSummary.py
# 2つのSummaryExcelファイルのグラフ系列をマージするプログラム
# =============================================================================

import copy
import shutil
import os
import re
import openpyxl

# =============================================================================
# ★★★ 設定エリア ★★★
# マージするファイルと系列名に付ける識別子を設定してください
# =============================================================================

# ファイル1の設定（Afterファイル）
FILE1_PATH = r"C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\40167\40167 Summary_After Rework .xlsx"
FILE1_SUFFIX = "After"   # ← 系列名に付ける識別子（例: "After" → "0dB After"）

# ファイル2の設定（Beforeファイル）
FILE2_PATH = r"C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\40167\40167 Summary_XVDAC before tripler rework.xlsx"
FILE2_SUFFIX = "Before"  # ← 系列名に付ける識別子（例: "Before" → "0dB Before"）

# 出力ファイルの保存先フォルダ（空欄の場合はFILE1_PATHと同じフォルダ）
OUTPUT_DIR = ""

# =============================================================================
# ★★★ 設定エリア ここまで ★★★
# =============================================================================

FILE2_SHEET_PREFIX = "_B_"


def get_series_title(ser):
    try:
        if ser.title is None:
            return None
        if hasattr(ser.title, 'v') and ser.title.v is not None:
            return ser.title.v
        if hasattr(ser.title, 'strRef') and ser.title.strRef:
            return ser.title.strRef.f
        return str(ser.title)
    except Exception:
        return None


def set_series_title(ser, new_title):
    if ser.title is not None and hasattr(ser.title, 'v') and ser.title.v is not None:
        ser.title.v = new_title
    elif ser.title is not None and hasattr(ser.title, 'strRef') and ser.title.strRef:
        ser.title.strRef.f = None
        ser.title.v = new_title
    else:
        from openpyxl.chart.series import SeriesLabel
        ser.title = SeriesLabel(v=new_title)


def copy_sheet_values(ws_src, ws_dst):
    """data_only=Trueで読んだシートの計算値をコピーする"""
    for row in ws_src.iter_rows():
        for cell in row:
            if cell.value is not None:
                ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)


def update_series_sheet_refs(ser, sheet_name_map):
    """系列内のシート参照をsheet_name_mapに従って書き換え、キャッシュをクリアする"""
    def remap(formula):
        if not formula:
            return formula
        for old_name, new_name in sheet_name_map.items():
            formula = formula.replace(f"'{old_name}'!", f"'{new_name}'!")
            formula = re.sub(
                rf"(?<!['\w])({re.escape(old_name)})!",
                f"'{new_name}'!",
                formula
            )
        return formula

    for attr in ['xVal', 'yVal', 'val']:
        ds = getattr(ser, attr, None)
        if ds and hasattr(ds, 'numRef') and ds.numRef and ds.numRef.f:
            ds.numRef.f = remap(ds.numRef.f)
            ds.numRef.numCache = None  # キャッシュクリア → Excelが再計算


def merge_summary(file1_path, file1_suffix, file2_path, file2_suffix, output_dir=""):
    """2つのExcelファイルのグラフ系列をマージする"""

    base_dir = output_dir if output_dir else os.path.dirname(file1_path)
    f1_name = os.path.splitext(os.path.basename(file1_path))[0]
    f2_name = os.path.splitext(os.path.basename(file2_path))[0]
    output_filename = f"Compare_{f1_name}_vs_{f2_name}.xlsx"
    output_path = os.path.join(base_dir, output_filename)

    print(f"マージ開始...")
    print(f"  File1 ({file1_suffix}): {os.path.basename(file1_path)}")
    print(f"  File2 ({file2_suffix}): {os.path.basename(file2_path)}")
    print(f"  出力: {output_filename}")
    print()

    # File1をベースとしてコピー
    shutil.copy2(file1_path, output_path)

    # マージファイル（File1ベース）を開く
    wb_merged = openpyxl.load_workbook(output_path)

    # File2: チャート参照用（系列オブジェクト取得）
    wb2_charts = openpyxl.load_workbook(file2_path)
    # File2: データ取得用（計算済み値）
    wb2_data = openpyxl.load_workbook(file2_path, data_only=True)

    # グラフのある共通シートを特定
    chart_sheets = [s for s in wb_merged.sheetnames
                    if s in wb2_charts.sheetnames and wb_merged[s]._charts]

    # File2の系列が参照するシートを収集
    file2_ref_sheets = set()
    for sname in chart_sheets:
        ws2 = wb2_charts[sname]
        for chart in ws2._charts:
            for ser in chart.series:
                for attr in ['xVal', 'yVal', 'val']:
                    ds = getattr(ser, attr, None)
                    if ds and hasattr(ds, 'numRef') and ds.numRef and ds.numRef.f:
                        raw = ds.numRef.f.split('!')[0].strip("'")
                        file2_ref_sheets.add(raw)

    # --- Step 1: File2のデータシートを計算済み値でマージファイルにコピー ---
    sheet_name_map = {}  # {元のシート名: 新しいシート名}
    for sname in sorted(file2_ref_sheets):
        if sname not in wb2_data.sheetnames:
            print(f"  警告: '{sname}' がFile2に見つかりません。スキップします。")
            continue
        new_name = f"{FILE2_SHEET_PREFIX}{sname}"[:31]
        sheet_name_map[sname] = new_name
        ws_src = wb2_data[sname]
        ws_dst = wb_merged.create_sheet(title=new_name)
        copy_sheet_values(ws_src, ws_dst)
        # コピーされたセル数を確認
        count = sum(1 for row in ws_dst.iter_rows() for c in row if c.value is not None)
        print(f"  データシートコピー: '{sname}' → '{new_name}' ({count} セル)")

    print()

    # --- Step 2: チャート系列のマージ ---
    total_charts = 0
    total_series_renamed = 0
    total_series_added = 0

    for sheet_name in chart_sheets:
        ws_merged = wb_merged[sheet_name]
        ws2 = wb2_charts[sheet_name]

        print(f"  シート '{sheet_name}': {len(ws_merged._charts)} チャート処理中...")

        for chart_idx, chart_merged in enumerate(ws_merged._charts):
            total_charts += 1

            # File1系列名に suffix1 を付加
            for ser in chart_merged.series:
                original_title = get_series_title(ser)
                if original_title is not None:
                    set_series_title(ser, f"{original_title} {file1_suffix}")
                    total_series_renamed += 1

            # File2系列をコピー・参照更新・suffix2 付加して追加
            if chart_idx < len(ws2._charts):
                chart2 = ws2._charts[chart_idx]
                for ser2 in chart2.series:
                    ser_copy = copy.deepcopy(ser2)
                    update_series_sheet_refs(ser_copy, sheet_name_map)
                    original_title2 = get_series_title(ser_copy)
                    if original_title2 is not None:
                        set_series_title(ser_copy, f"{original_title2} {file2_suffix}")
                    chart_merged.series.append(ser_copy)
                    total_series_added += 1

    wb_merged.save(output_path)
    wb2_charts.close()
    wb2_data.close()

    print()
    print(f"マージ完了!")
    print(f"  処理チャート数    : {total_charts}")
    print(f"  リネーム系列数    : {total_series_renamed} ({file1_suffix}付加)")
    print(f"  追加系列数        : {total_series_added} ({file2_suffix}付加)")
    print(f"  コピーシート数    : {len(sheet_name_map)} (File2データ用)")
    print(f"  保存先            : {output_path}")

    return output_path


if __name__ == "__main__":
    merge_summary(
        file1_path=FILE1_PATH,
        file1_suffix=FILE1_SUFFIX,
        file2_path=FILE2_PATH,
        file2_suffix=FILE2_SUFFIX,
        output_dir=OUTPUT_DIR,
    )
