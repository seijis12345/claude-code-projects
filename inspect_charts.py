from openpyxl import load_workbook
import os

target_sheets = ["summary P1/P2", "Compare P1/P2"]

def get_chart_title(chart):
    title = chart.title
    if title is None:
        return None
    if isinstance(title, str):
        return title
    try:
        # Title object with rich text
        paragraphs = title.tx.rich.p
        texts = []
        for p in paragraphs:
            for r in p.r:
                texts.append(r.t)
        return "".join(texts)
    except:
        pass
    try:
        return title.tx.strRef.f
    except:
        pass
    return str(type(title))

# Check first file that has target sheets
test_file = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics\10259\10259SummaryXVDAC.xlsx"
wb = load_workbook(test_file)
print(f"File: {os.path.basename(test_file)}")
print(f"Sheets: {wb.sheetnames}")

for sheet_name in target_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        charts = ws._charts
        print(f"\nSheet '{sheet_name}': {len(charts)} charts")
        for i, chart in enumerate(charts):
            title = get_chart_title(chart)
            print(f"  Chart {i+1} (index {i}): {repr(title)}")
