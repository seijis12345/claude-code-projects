"""
SummaryXVDAC ファイルを2つ統合して Compare ファイルを作成するスクリプト

系列名:
  元データ (FILE_A) → "LABEL_A 0dB" ... "LABEL_A 60dB"
  追加データ (FILE_B) → "LABEL_B 0dB" ... "LABEL_B 60dB"
"""

import win32com.client
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# ════════════════════════════════════════════════════════════════
#  ▼▼▼  ここを変更してください  ▼▼▼
FILE_A  = r'C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\10259\10259 Summary_after swap tripler.xlsx'
FILE_B  = r'C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\39992\Swap_Tripler\39992 Summary_after swap tripler.xlsx'
LABEL_A = '10259_After'
LABEL_B = '39992_After'
OUT_FILE = r'C:\Users\seijis\OneDrive - Keysight Technologies\Diva50_MaxPower2026\01_Kobe_Trip_Penang_Data\04_MMA_Data_Pmax_Harmonics\Compare_10259_39992_After.xlsx'
# ▲▲▲  ここを変更してください  ▲▲▲
# ════════════════════════════════════════════════════════════════

# ── 行範囲定義 ────────────────────────────────────────────────────
R_START     = 5
R_END       = 137
R_DELTA_END = 121   # Chart2 (Delta/ref60dB) の行末

# ── 元ファイルの列グループ (1-indexed) ───────────────────────────
# 各グループ: (元列開始, 元列終了, データ行末)
ORIG_GROUPS = [
    (46,  58,  R_END),       # AT:BF  P1Max       → Chart1
    (61,  73,  R_DELTA_END), # BI:BU  Delta       → Chart2
    (76,  88,  R_END),       # BX:CJ  1/3th       → Chart3
    (91,  103, R_END),       # CM:CY  1/3 dBm     → Chart7
    (106, 118, R_END),       # DB:DN  2/3th       → Chart4
    (121, 133, R_END),       # DQ:EC  2/3 dBm     → Chart8
]
COL_FREQ_ORIG = 45  # AS: X軸周波数

# ── FILE_B データを書き込む新列 (連続配置) ───────────────────────
NEW_COL_BASE = 135  # EG から開始
COL_FREQ_NEW = NEW_COL_BASE  # EG(135): 周波数X軸

# 元グループ → 新グループ のマッピングを構築
NEW_GROUPS = []
offset = 1
for (os_, oe_, re_) in ORIG_GROUPS:
    ns_ = NEW_COL_BASE + offset
    ne_ = ns_ + (oe_ - os_)
    NEW_GROUPS.append((ns_, ne_, re_))
    offset += (oe_ - os_ + 1)

# 元列 → 新列 の変換関数
def map_col(orig_col):
    for (os_, oe_, _), (ns_, ne_, _) in zip(ORIG_GROUPS, NEW_GROUPS):
        if os_ <= orig_col <= oe_:
            return ns_ + (orig_col - os_)
    return None


def cl(col):
    return get_column_letter(col)


def read_values(filepath, sheet_name):
    """openpyxl data_only で計算済み値を辞書 {(row,col): value} で取得"""
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    data = {}
    for r in range(R_START, R_END + 1):
        # 周波数列
        v = ws.cell(r, COL_FREQ_ORIG).value
        if v is not None:
            data[(r, COL_FREQ_ORIG)] = v
        # 各グループ列
        for (os_, oe_, re_) in ORIG_GROUPS:
            if r > re_:
                continue
            for c in range(os_, oe_ + 1):
                v = ws.cell(r, c).value
                if v is not None:
                    data[(r, c)] = v
    return data


def write_a_values_to_sheet(com_sheet, data_a):
    """FILE_A 計算済み値を元の列位置に上書き（数式を値に置換）"""
    # 周波数列
    vals = [(data_a.get((r, COL_FREQ_ORIG), ''),) for r in range(R_START, R_END + 1)]
    com_sheet.Range(
        com_sheet.Cells(R_START, COL_FREQ_ORIG),
        com_sheet.Cells(R_END,   COL_FREQ_ORIG)
    ).Value = vals
    # 各グループ列
    for (os_, oe_, re_) in ORIG_GROUPS:
        for c in range(os_, oe_ + 1):
            vals = [(data_a.get((r, c), ''),) for r in range(R_START, re_ + 1)]
            com_sheet.Range(
                com_sheet.Cells(R_START, c),
                com_sheet.Cells(re_,     c)
            ).Value = vals


def write_b_to_sheet(com_sheet, data_b):
    """FILE_B データを Compare シートの新列に書き込む"""
    # 周波数列
    vals = [(data_b.get((r, COL_FREQ_ORIG), ''),) for r in range(R_START, R_END + 1)]
    com_sheet.Range(
        com_sheet.Cells(R_START, COL_FREQ_NEW),
        com_sheet.Cells(R_END,   COL_FREQ_NEW)
    ).Value = vals

    # 各グループ
    for (os_, oe_, re_), (ns_, ne_, _) in zip(ORIG_GROUPS, NEW_GROUPS):
        for src_c, dst_c in zip(range(os_, oe_ + 1), range(ns_, ne_ + 1)):
            vals = [(data_b.get((r, src_c), ''),) for r in range(R_START, re_ + 1)]
            com_sheet.Range(
                com_sheet.Cells(R_START, dst_c),
                com_sheet.Cells(re_,     dst_c)
            ).Value = vals


def parse_series_refs(formula):
    """
    SERIES式から X・Y 参照情報を取得する
    Returns: (x_col_s, x_col_e, x_row_s, x_row_e,
              y_col_s, y_col_e, y_row_s, y_row_e)
    """
    refs = re.findall(r'\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)', formula)
    if len(refs) < 2:
        return None
    xcs, xrs, xce, xre = refs[0]
    ycs, yrs, yce, yre = refs[1]
    return (column_index_from_string(xcs), column_index_from_string(xce),
            int(xrs), int(xre),
            column_index_from_string(ycs), column_index_from_string(yce),
            int(yrs), int(yre))


def make_series_formula(name, sheet, x_cs, x_ce, x_rs, x_re, y_cs, y_ce, y_rs, y_re, order):
    """SERIES式文字列を生成"""
    x_ref = f"'{sheet}'!${cl(x_cs)}${x_rs}:${cl(x_ce)}${x_re}"
    y_ref = f"'{sheet}'!${cl(y_cs)}${y_rs}:${cl(y_ce)}${y_re}"
    return f'=SERIES("{name}",{x_ref},{y_ref},{order})'


def update_charts(com_sheet, sheet_name, label_a, label_b, excel_app):
    n = com_sheet.ChartObjects().Count
    print(f'  {sheet_name}: チャート数 = {n}')

    for i in range(1, n + 1):
        try:
            ch = com_sheet.ChartObjects(i).Chart
        except Exception as e:
            print(f'  Chart {i}: 取得失敗 → スキップ ({e})')
            continue

        ns = ch.SeriesCollection().Count
        if ns == 0:
            continue

        # 既存系列の情報を保存
        orig = []
        for j in range(1, ns + 1):
            s = ch.SeriesCollection(j)
            info = parse_series_refs(s.Formula)
            orig.append((s.Name, info))

        # ── Step1: 既存系列をリネーム ──────────────────────────
        for j in range(1, ns + 1):
            ch.SeriesCollection(j).Name = f'{label_a} {ch.SeriesCollection(j).Name}'

        # ── Step2: FILE_B 系列を追加 ──────────────────────────
        added = 0
        for j, (name, info) in enumerate(orig):
            if info is None:
                continue
            x_cs, x_ce, x_rs, x_re, y_cs, y_ce, y_rs, y_re = info

            if x_cs == COL_FREQ_ORIG:
                new_x_cs = COL_FREQ_NEW
                new_x_ce = COL_FREQ_NEW
            else:
                new_x_cs = x_cs
                new_x_ce = x_ce

            new_y_cs = map_col(y_cs)
            new_y_ce = map_col(y_ce)
            if new_y_cs is None or new_y_ce is None:
                continue

            formula = make_series_formula(
                f'{label_b} {name}', sheet_name,
                new_x_cs, new_x_ce, x_rs, x_re,
                new_y_cs, new_y_ce, y_rs, y_re,
                ns + added + 1
            )
            try:
                new_s = ch.SeriesCollection().NewSeries()
                new_s.Formula = formula
                added += 1
            except Exception as e:
                # Excel が不安定な場合は少し待ってリトライ
                import time
                time.sleep(0.5)
                try:
                    new_s = ch.SeriesCollection().NewSeries()
                    new_s.Formula = formula
                    added += 1
                except Exception as e2:
                    print(f'  Chart {i} S{j+1} 追加失敗(リトライ後): {e2}')

        print(f'  Chart {i}: {ns}系列リネーム + {added}系列追加 完了')


# ══════════════════════════════════════════════════════════════
#  メイン
# ══════════════════════════════════════════════════════════════
print('=== データ読み込み (data_only) ===')
data_a_p1 = read_values(FILE_A, 'summary P1')
data_a_p2 = read_values(FILE_A, 'summary P2')
data_b_p1 = read_values(FILE_B, 'summary P1')
data_b_p2 = read_values(FILE_B, 'summary P2')
print(f'  FILE_A summary P1: {len(data_a_p1)} セル')
print(f'  FILE_A summary P2: {len(data_a_p2)} セル')
print(f'  FILE_B summary P1: {len(data_b_p1)} セル')
print(f'  FILE_B summary P2: {len(data_b_p2)} セル')

excel = win32com.client.Dispatch('Excel.Application')
excel.Visible          = False
excel.DisplayAlerts    = False
excel.ScreenUpdating   = False
excel.EnableEvents     = False

try:
    wb_a = excel.Workbooks.Open(os.path.abspath(FILE_A))
    wb_new = excel.Workbooks.Add()
    excel.Calculation  = -4135   # xlCalculationManual (ワークブックオープン後に設定)

    # デフォルトシートを1枚残して削除
    while wb_new.Sheets.Count > 1:
        wb_new.Sheets(wb_new.Sheets.Count).Delete()
    wb_new.Sheets(1).Name = '_tmp'

    # summary P2 → Compare P2 (先にコピー)
    wb_a.Sheets('summary P2').Copy(Before=wb_new.Sheets('_tmp'))
    compare_p2 = excel.ActiveSheet
    compare_p2.Name = 'Compare P2'

    # summary P1 → Compare P1 (Compare P2 の前)
    wb_a.Sheets('summary P1').Copy(Before=compare_p2)
    compare_p1 = excel.ActiveSheet
    compare_p1.Name = 'Compare P1'

    wb_new.Sheets('_tmp').Delete()

    # ── Compare P1 処理 ──────────────────────────────────────
    print('\n=== Compare P1 処理 ===')
    write_a_values_to_sheet(compare_p1, data_a_p1)
    print('  FILE_A 計算済み値を上書き完了')
    write_b_to_sheet(compare_p1, data_b_p1)
    print('  FILE_B データ書き込み完了')
    update_charts(compare_p1, 'Compare P1', LABEL_A, LABEL_B, excel)

    # ── Compare P2 処理 ──────────────────────────────────────
    print('\n=== Compare P2 処理 ===')
    write_a_values_to_sheet(compare_p2, data_a_p2)
    print('  FILE_A 計算済み値を上書き完了')
    write_b_to_sheet(compare_p2, data_b_p2)
    print('  FILE_B データ書き込み完了')
    update_charts(compare_p2, 'Compare P2', LABEL_A, LABEL_B, excel)

    # 保存
    wb_new.SaveAs(os.path.abspath(OUT_FILE), FileFormat=51)
    print(f'\n保存完了: {OUT_FILE}')

finally:
    try: wb_a.Close(SaveChanges=False)
    except: pass
    try: wb_new.Close(SaveChanges=False)
    except: pass
    excel.Quit()

print('完了しました。')
