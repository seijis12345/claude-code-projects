from openpyxl import load_workbook
import glob, os

folder = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics"
target_sheets = ["summary P1/P2", "Compare P1/P2"]
files = glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True)

for f in files:
    try:
        wb = load_workbook(f, read_only=True)
        found = [s for s in target_sheets if s in wb.sheetnames]
        if found:
            print(f"{os.path.basename(f)}: {found}")
        wb.close()
    except Exception as e:
        print(f"ERROR {os.path.basename(f)}: {e}")
