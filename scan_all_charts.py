from openpyxl import load_workbook
import glob, os

folder = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics"
target_sheets = ["summary P1", "summary P2", "Compare P1", "Compare P2"]

def get_chart_title(chart):
    title = chart.title
    if title is None:
        return None
    if isinstance(title, str):
        return title
    try:
        paragraphs = title.tx.rich.p
        texts = []
        for p in paragraphs:
            for r in p.r:
                texts.append(r.t)
        return "".join(texts)
    except:
        pass
    return repr(title)

files = glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True)

for f in files:
    try:
        wb = load_workbook(f)
        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            charts = ws._charts
            if len(charts) >= 7:
                t7 = get_chart_title(charts[6])
                if t7 and '1/3 sub-harmonic' in t7 and '2/3' not in t7:
                    print(f"[Chart7 NEEDS FIX] {os.path.basename(f)} | {sheet_name} | {repr(t7)}")
            if len(charts) >= 4:
                t4 = get_chart_title(charts[3])
                if t4 and '2/3th sub-harmonics' in t4:
                    print(f"[Chart4 NEEDS FIX] {os.path.basename(f)} | {sheet_name} | {repr(t4)}")
        wb.close()
    except Exception as e:
        print(f"ERROR {os.path.basename(f)}: {e}")

print("Scan complete.")
