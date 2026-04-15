"""
Fix chart titles in Excel files:
  - summary P1, summary P2, Compare P1, Compare P2 sheets
  - Chart 7 (index 6): '1/3 sub-harmonic' → '2/3 sub-harmonic'
  - Chart 4 (index 3): '2/3th sub-harmonics' → '1/3th sub-harmonics'
"""
from openpyxl import load_workbook
import glob, os

FOLDER = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics"
TARGET_SHEETS = ["summary P1", "summary P2", "Compare P1", "Compare P2"]


def get_title_runs(chart):
    title = chart.title
    if title is None or isinstance(title, str):
        return []
    try:
        runs = []
        for p in title.tx.rich.p:
            for r in p.r:
                runs.append(r)
        return runs
    except Exception:
        return []


def get_chart_title_text(chart):
    runs = get_title_runs(chart)
    if runs:
        return "".join(r.t for r in runs)
    if isinstance(chart.title, str):
        return chart.title
    return None


def replace_in_title(chart, old, new):
    """Replace old substring with new in chart title runs. Returns True if changed."""
    runs = get_title_runs(chart)
    if not runs:
        return False
    full = "".join(r.t for r in runs)
    if old not in full:
        return False
    # Replace within individual runs first
    for r in runs:
        if old in r.t:
            r.t = r.t.replace(old, new)
            return True
    # Fallback: text spans multiple runs - rewrite all into first run
    new_full = full.replace(old, new)
    runs[0].t = new_full
    for r in runs[1:]:
        r.t = ""
    return True


files = glob.glob(os.path.join(FOLDER, "**", "*.xlsx"), recursive=True)
print(f"Scanning {len(files)} xlsx files...\n")

total_changes = 0

for f in sorted(files):
    try:
        wb = load_workbook(f)
        file_changed = False

        for sheet_name in TARGET_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            charts = wb[sheet_name]._charts

            # Chart 7 (index 6): '1/3 sub-harmonic' → '2/3 sub-harmonic'
            if len(charts) >= 7:
                t = get_chart_title_text(charts[6])
                if t and '1/3 sub-harmonic' in t and '2/3' not in t:
                    old_t = t
                    if replace_in_title(charts[6], '1/3 sub-harmonic', '2/3 sub-harmonic'):
                        new_t = get_chart_title_text(charts[6])
                        print(f"[Chart7] {os.path.basename(f)} | {sheet_name}")
                        print(f"  Before: {repr(old_t)}")
                        print(f"  After:  {repr(new_t)}")
                        file_changed = True
                        total_changes += 1

            # Chart 4 (index 3): '2/3th sub-harmonics' → '1/3th sub-harmonics'
            if len(charts) >= 4:
                t = get_chart_title_text(charts[3])
                if t and '2/3th sub-harmonics' in t:
                    old_t = t
                    if replace_in_title(charts[3], '2/3th sub-harmonics', '1/3th sub-harmonics'):
                        new_t = get_chart_title_text(charts[3])
                        print(f"[Chart4] {os.path.basename(f)} | {sheet_name}")
                        print(f"  Before: {repr(old_t)}")
                        print(f"  After:  {repr(new_t)}")
                        file_changed = True
                        total_changes += 1

        if file_changed:
            wb.save(f)
            print(f"  => Saved: {os.path.relpath(f, FOLDER)}\n")
        wb.close()

    except Exception as e:
        print(f"ERROR {os.path.basename(f)}: {e}")

print(f"Done. Total changes: {total_changes}")
