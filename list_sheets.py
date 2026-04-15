from openpyxl import load_workbook
import glob, os

folder = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics"
files = glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True)

shown = set()
for f in files[:10]:
    try:
        wb = load_workbook(f, read_only=True)
        name = os.path.basename(f)
        print(f"{name}: {wb.sheetnames}")
        wb.close()
    except Exception as e:
        print(f"ERROR {os.path.basename(f)}: {e}")
