from openpyxl import load_workbook

def get_chart_title(chart):
    title = chart.title
    if title is None:
        return None
    if isinstance(title, str):
        return title
    try:
        paragraphs = title.tx.rich.p
        texts = []
        for p in paragraphs:
            for r in p.r:
                texts.append(r.t)
        return "".join(texts)
    except:
        pass
    return repr(title)

test_file = r"C:\Users\seijis\Desktop\04_MMA_Data_Pmax_Harmonics\10259\10259SummaryXVDAC.xlsx"
wb = load_workbook(test_file)

for sheet_name in ["summary P1", "summary P2"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        charts = ws._charts
        print(f"\nSheet '{sheet_name}': {len(charts)} charts")
        for i, chart in enumerate(charts):
            title = get_chart_title(chart)
            print(f"  Chart {i+1}: {repr(title)}")
